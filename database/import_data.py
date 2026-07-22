"""
冶金平台 v2 — 数据导入脚本

从 绿色低碳冶金_数据库资料源与120小模型清单.xlsx 提取数据，
导入本地 PostgreSQL metallurgy 库的 metallurgy_v2 schema。

用法:
  python import_data.py

前置条件:
  - PostgreSQL 运行在 127.0.0.1:5432 (trust auth)
  - metallurgy 库已创建，建表 SQL 已执行
  - 依赖: pip install openpyxl psycopg2-binary
"""
import sys
import os
from datetime import datetime

# 添加 Tools 和 database 目录到路径
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_PROJECT_DIR = os.path.dirname(_SCRIPT_DIR)
_TOOLS_DIR = os.path.join(_PROJECT_DIR, 'Tools')
sys.path.insert(0, _TOOLS_DIR)
sys.path.insert(0, _PROJECT_DIR)
import openpyxl
import psycopg2
from psycopg2.extras import execute_values

# ── 数据库连接 ──
DB_CONFIG = {
    'host': '127.0.0.1',
    'port': 5432,
    'database': 'metallurgy',
    'user': 'postgres',
    'password': '',
    'connect_timeout': 5,
}

EXCEL_PATH = os.path.join(
    os.path.dirname(os.path.dirname(__file__)),
    '绿色低碳冶金_数据库资料源与120小模型清单.xlsx'
)

# ── 内置热化学数据（10种反应）──
BUILTIN_THERMO_DATA = [
    # (reaction_id, species, property_type, temperature, value, unit, source, dataset_id)
    # ΔHf° data at 298K from NIST-JANAF
    ('R001', 'CO2(g)', '生成焓_ΔHf°', 298, -393.5, 'kJ/mol', 'NIST-JANAF', 'DS002'),
    ('R001', 'CO2(g)', '生成焓_ΔHf°', 500, -393.5, 'kJ/mol', 'NIST-JANAF', 'DS002'),
    ('R002', 'CO(g)', '生成焓_ΔHf°', 298, -110.5, 'kJ/mol', 'NIST-JANAF', 'DS002'),
    ('R003', 'FeO(s)', '生成焓_ΔHf°', 298, -272.0, 'kJ/mol', 'NIST-JANAF', 'DS002'),
    ('R003', 'Fe(s)', '生成焓_ΔHf°', 298, 0.0, 'kJ/mol', 'NIST-JANAF', 'DS002'),
    ('R004', 'Fe2O3(s)', '生成焓_ΔHf°', 298, -824.2, 'kJ/mol', 'NIST-JANAF', 'DS002'),
    ('R005', 'Fe3O4(s)', '生成焓_ΔHf°', 298, -1118.4, 'kJ/mol', 'NIST-JANAF', 'DS002'),
    ('R006', 'CaCO3(s)', '生成焓_ΔHf°', 298, -1207.6, 'kJ/mol', 'NIST-JANAF', 'DS002'),
    ('R006', 'CaO(s)', '生成焓_ΔHf°', 298, -634.9, 'kJ/mol', 'NIST-JANAF', 'DS002'),
    ('R007', 'SiO2(s)', '生成焓_ΔHf°', 298, -910.7, 'kJ/mol', 'NIST-JANAF', 'DS002'),
    ('R008', 'Si(s)', '生成焓_ΔHf°', 298, 0.0, 'kJ/mol', 'NIST-JANAF', 'DS002'),
    ('R009', 'MnO(s)', '生成焓_ΔHf°', 298, -385.2, 'kJ/mol', 'NIST-JANAF', 'DS002'),
    ('R010', 'Al2O3(s)', '生成焓_ΔHf°', 298, -1675.7, 'kJ/mol', 'NIST-JANAF', 'DS002'),
    ('R010', 'Al(s)', '生成焓_ΔHf°', 298, 0.0, 'kJ/mol', 'NIST-JANAF', 'DS002'),
]

# ── 内置反应定义 ──
BUILTIN_REACTIONS = [
    ('R001', 'C + O₂ → CO₂', '碳完全燃烧', '氧化',
     '[{"species":"C","coeff":1,"phase":"s"},{"species":"O2","coeff":1,"phase":"g"}]',
     '[{"species":"CO2","coeff":1,"phase":"g"}]'),
    ('R002', '2C + O₂ → 2CO', '碳不完全燃烧', '氧化',
     '[{"species":"C","coeff":2,"phase":"s"},{"species":"O2","coeff":1,"phase":"g"}]',
     '[{"species":"CO","coeff":2,"phase":"g"}]'),
    ('R003', 'FeO + C → Fe + CO', '氧化亚铁碳还原', '还原',
     '[{"species":"FeO","coeff":1,"phase":"s"},{"species":"C","coeff":1,"phase":"s"}]',
     '[{"species":"Fe","coeff":1,"phase":"s"},{"species":"CO","coeff":1,"phase":"g"}]'),
    ('R004', 'Fe₂O₃ + 3CO → 2Fe + 3CO₂', '赤铁矿间接还原', '还原',
     '[{"species":"Fe2O3","coeff":1,"phase":"s"},{"species":"CO","coeff":3,"phase":"g"}]',
     '[{"species":"Fe","coeff":2,"phase":"s"},{"species":"CO2","coeff":3,"phase":"g"}]'),
    ('R005', 'Fe₃O₄ + 4CO → 3Fe + 4CO₂', '磁铁矿间接还原', '还原',
     '[{"species":"Fe3O4","coeff":1,"phase":"s"},{"species":"CO","coeff":4,"phase":"g"}]',
     '[{"species":"Fe","coeff":3,"phase":"s"},{"species":"CO2","coeff":4,"phase":"g"}]'),
    ('R006', 'CaCO₃ → CaO + CO₂', '碳酸钙分解', '分解',
     '[{"species":"CaCO3","coeff":1,"phase":"s"}]',
     '[{"species":"CaO","coeff":1,"phase":"s"},{"species":"CO2","coeff":1,"phase":"g"}]'),
    ('R007', 'SiO₂ + 2C → Si + 2CO', '二氧化硅碳还原', '还原',
     '[{"species":"SiO2","coeff":1,"phase":"s"},{"species":"C","coeff":2,"phase":"s"}]',
     '[{"species":"Si","coeff":1,"phase":"s"},{"species":"CO","coeff":2,"phase":"g"}]'),
    ('R008', '2FeO + Si → 2Fe + SiO₂', '硅还原氧化亚铁', '还原',
     '[{"species":"FeO","coeff":2,"phase":"s"},{"species":"Si","coeff":1,"phase":"s"}]',
     '[{"species":"Fe","coeff":2,"phase":"s"},{"species":"SiO2","coeff":1,"phase":"s"}]'),
    ('R009', 'MnO + C → Mn + CO', '氧化锰碳还原', '还原',
     '[{"species":"MnO","coeff":1,"phase":"s"},{"species":"C","coeff":1,"phase":"s"}]',
     '[{"species":"Mn","coeff":1,"phase":"s"},{"species":"CO","coeff":1,"phase":"g"}]'),
    ('R010', 'Fe₂O₃ + 2Al → 2Fe + Al₂O₃', '铝热反应', '置换',
     '[{"species":"Fe2O3","coeff":1,"phase":"s"},{"species":"Al","coeff":2,"phase":"s"}]',
     '[{"species":"Fe","coeff":2,"phase":"s"},{"species":"Al2O3","coeff":1,"phase":"s"}]'),
]


def connect():
    """连接 PostgreSQL"""
    conn = psycopg2.connect(**DB_CONFIG)
    conn.set_session(autocommit=True)
    return conn


def import_dataset_registry(conn):
    """从 Excel 导入数据源目录"""
    print('📖 读取数据源...')
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['01_数据库资料源']

    rows = []
    for row_idx in range(2, ws.max_row + 1):
        ds_id = ws.cell(row=row_idx, column=1).value
        if not ds_id:
            continue
        rows.append((
            ds_id,
            ws.cell(row=row_idx, column=3).value or '',       # name
            ws.cell(row=row_idx, column=2).value or '',       # category
            ws.cell(row=row_idx, column=4).value or '',       # provider
            ws.cell(row=row_idx, column=7).value or '公开',    # license
            ws.cell(row=row_idx, column=12).value or '',       # access_url
            ws.cell(row=row_idx, column=6).value or '',        # ingestion_mode
            '1.0' if ds_id else '',                            # version
            datetime.now(),                                     # retrieved_at
            None,                                               # checksum
            '数据组',                                           # owner
            '公开' if '公开' in str(ws.cell(row=row_idx, column=7).value or '') else '内部',  # security_level
            'A',                                                # quality_grade
            None,                                               # lineage_json
        ))

    sql = """INSERT INTO metallurgy_v2.dataset_registry
        (dataset_id, name, category, provider, license, access_url,
         ingestion_mode, version, retrieved_at, checksum, owner,
         security_level, quality_grade, lineage_json)
        VALUES %s
        ON CONFLICT (dataset_id) DO UPDATE SET
            name = EXCLUDED.name,
            category = EXCLUDED.category,
            provider = EXCLUDED.provider,
            license = EXCLUDED.license,
            access_url = EXCLUDED.access_url,
            ingestion_mode = EXCLUDED.ingestion_mode,
            version = EXCLUDED.version
    """
    execute_values(conn.cursor(), sql, rows)
    print(f'  ✅ 导入 {len(rows)} 个数据源')
    wb.close()


def import_model_registry(conn):
    """注册现有模型到 model_registry"""
    print('  ⚠️ 跳过模型注册（需在 Tools 目录下执行）')
    return

    rows = []
    for entry in reg.list_models():
        rows.append((
            entry['model_id'],
            entry['name'],
            entry['scenario'],
            entry['model_type'],
            entry['api_name'],
            str(entry['input_schema_json']),
            str(entry['output_schema_json']),
            entry['applicable_boundary'],
            str(entry.get('validation_rules', [])),
            entry['priority'],
            '开发组',
            entry.get('status', 'dev'),
        ))

    sql = """INSERT INTO metallurgy_v2.model_registry
        (model_id, name, scenario, model_type, api_name,
         input_schema_json, output_schema_json, applicable_boundary,
         validation_rules, priority, owner, status)
        VALUES %s
        ON CONFLICT (model_id) DO UPDATE SET
            name = EXCLUDED.name,
            status = EXCLUDED.status,
            priority = EXCLUDED.priority
    """
    execute_values(conn.cursor(), sql, rows)
    print(f'  ✅ 注册 {len(rows)} 个模型')
    print(f'     模型: {[r[0] for r in rows]}')

    # 添加版本记录
    version_rows = [(r[0], '1.0.0', f'code://models_core.{r[4]}', None, None, None, 'requirements.txt', datetime.now(), None, '初始版本') for r in rows]
    v_sql = """INSERT INTO metallurgy_v2.model_versions
        (model_id, version, artifact_uri, git_commit, data_snapshot_id,
         parameter_json, dependency_lock, created_at, approved_by, release_notes)
        VALUES %s
        ON CONFLICT (model_id, version) DO NOTHING
    """
    execute_values(conn.cursor(), v_sql, version_rows)
    print(f'  ✅ 添加 {len(version_rows)} 条版本记录')


def import_reactions(conn):
    """导入内置反应定义"""
    sql = """INSERT INTO metallurgy_v2.reaction_definition
        (reaction_id, reaction_equation, name, category, reactants, products)
        VALUES %s
        ON CONFLICT (reaction_id) DO UPDATE SET
            name = EXCLUDED.name
    """
    execute_values(conn.cursor(), sql, BUILTIN_REACTIONS)
    print(f'  ✅ 导入 {len(BUILTIN_REACTIONS)} 条反应定义')

    # 导入 ΔHf° 数据
    tp_sql = """INSERT INTO metallurgy_v2.thermodynamic_property
        (dataset_id, species, property_type, temperature, value, unit, source_ref, quality_grade)
        VALUES %s
    """
    tp_rows = [(d[4], d[1], d[2], d[3], d[4], d[5], d[6], 'A') for d in BUILTIN_THERMO_DATA]
    tp_rows_fixed = []
    for d in BUILTIN_THERMO_DATA:
        tp_rows_fixed.append((d[7], d[1], d[2], float(d[3]), float(d[4]), d[5], d[6], 'A'))
    execute_values(conn.cursor(), tp_sql, tp_rows_fixed)
    print(f'  ✅ 导入 {len(BUILTIN_THERMO_DATA)} 条热力学数据')


def import_model_calls_sample(conn):
    """插入几条示例调用日志"""
    logs = [
        ('TRACE-000001', 'CALL-000001', 'A002', '1.0.0', datetime.now(),
         '{"formula":"Fe2(SO4)3"}', '{"passed":true}', '{"elements":{"Fe":2,"S":3,"O":12},"molar_mass":399.858}',
         1.0, None, 35, 'success', None, 'system'),
        ('TRACE-000002', 'CALL-000002', 'B008', '1.0.0', datetime.now(),
         '{"reaction":"FeO + C → Fe + CO","temperature":1873}', '{"passed":true}',
         '{"delta_G":-122.95,"direction":"正向强烈自发"}',
         0.95, None, 12, 'success', None, 'user_001'),
        ('TRACE-000003', 'CALL-000003', 'A001', '1.0.0', datetime.now(),
         '{"value":100,"source_unit":"°C","target_unit":"K"}', '{"passed":true}',
         '{"value":373.15,"target_unit":"K"}',
         1.0, None, 5, 'success', None, 'user_001'),
    ]
    sql = """INSERT INTO metallurgy_v2.invocation_logs
        (trace_id, call_id, model_id, model_version, requested_at,
         input_json, boundary_check, output_json, confidence,
         validation_result, runtime_ms, status, error_code, user_or_agent)
        VALUES %s
    """
    execute_values(conn.cursor(), sql, logs)
    print(f'  ✅ 插入 {len(logs)} 条示例调用日志')


def main():
    print('=' * 56)
    print('  冶金平台 v2 — 数据导入')
    print('=' * 56)

    conn = connect()
    cur = conn.cursor()

    # 检查 metallurgy_v2 schema 是否存在
    cur.execute("SELECT EXISTS(SELECT 1 FROM information_schema.schemata WHERE schema_name = 'metallurgy_v2')")
    if not cur.fetchone()[0]:
        print('❌ metallurgy_v2 schema 不存在，请先执行建表 SQL')
        return

    print('\n📦 1. 数据源目录导入')
    import_dataset_registry(conn)

    print('\n🤖 2. 模型注册')
    try:
        import_model_registry(conn)
    except Exception as e:
        print(f'  ⚠️ 模型注册失败（可能是依赖问题）: {e}')

    print('\n⚗️ 3. 反应定义和热力学数据')
    import_reactions(conn)

    print('\n📝 4. 示例调用日志')
    try:
        import_model_calls_sample(conn)
    except Exception as e:
        print(f'  ⚠️ 示例日志插入失败: {e}')

    # 验证
    cur.execute("SELECT COUNT(*) FROM metallurgy_v2.dataset_registry")
    ds_count = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM metallurgy_v2.model_registry")
    m_count = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM metallurgy_v2.reaction_definition")
    r_count = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM metallurgy_v2.thermodynamic_property")
    tp_count = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM metallurgy_v2.invocation_logs")
    log_count = cur.fetchone()[0]

    print('\n' + '=' * 56)
    print('  📊 导入统计')
    print('=' * 56)
    print(f'  数据源:        {ds_count} 条')
    print(f'  模型注册:      {m_count} 条')
    print(f'  反应定义:      {r_count} 条')
    print(f'  热力学数据:    {tp_count} 条')
    print(f'  调用日志:      {log_count} 条')
    print('=' * 56)

    conn.close()
    print('\n✅ 导入完成！')


if __name__ == '__main__':
    main()
